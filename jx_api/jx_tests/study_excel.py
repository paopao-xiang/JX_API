# -*- coding:utf-8 _*-
""" 
@author:paopao
@time: 2020/11/10:11:44
@email:1332170414@qq.com
@function：
"""

import openpyxl
from jx_api.common import contants

class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """

    def __init__(self):
        self.phone=None
        self.data=None


class DoExcel:

    def __init__(self, file_name, sheet_name):
        # 异常处理
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row  # 获取最大行数

        cases = []  # 列表，存放所有的测试用例
        for r in range(2, max_row + 1):
            # case = {}
            # case['case_id'] = self.sheet.cell(row=r, column=1)
            # case['title'] = self.sheet.cell(row=r, column=2)
            case = Case()  # 实例
            case.phone = self.sheet.cell(row=r, column=1).value
            case.data=self.sheet.cell(row=r, column=2).value
            if case.data.find('case.phone') != -1:
                case.data = case.data.replace('case.phone', str(case.phone))
                self.sheet.cell(r, 2).value = case.data
        self.workbook.save(filename=self.file_name)
        self.workbook.close()
        return cases  # 返回case列表

if __name__ == '__main__':
    resp=DoExcel(contants.case_file,'phone')
    data=resp.get_cases()

